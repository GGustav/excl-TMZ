import openpyxl
from openpyxl import Workbook
from utility import *
import copy

# Sure wishing I had fuzzy search right about now to do this automatically from the sheet
alphatel_A = ['B Number', 'A Start Location Name',
              'B Start Location Name', 'A End Location Name', 'B End Location Name', 'A Start Location CGI', 'B Start Location CGI', 'A End Location CGI', 'B End Location CGI',
              'A Start Location Latitude', 'A Start Location Longitude', 'A Start Location Bearing', 'B Start Location Latitude', 'B Start Location Longitude', 'B Start Location Bearing',
              'A End Location Latitude', 'A End Location Longitude', 'A End Location Bearing', 'B End Location Latitude', 'B End Location Longitude', 'B End Location Bearing']

# Create alphatel_B automatically since I am not typing all that out again
alphatel_B = copy.deepcopy(alphatel_A)
for i in range(0, len(alphatel_A)):
    if alphatel_B[i][0] == 'A':
        alphatel_B[i] = alphatel_B[i].replace('A', 'B', 1)
    elif alphatel_B[i][0] == 'B':
        alphatel_B[i] = alphatel_B[i].replace('B', 'A', 1)

alpha_tz = pytz.timezone("Etc/UTC")

def alpha_sheet(original_sheet, targetnumber):
    original_headers = column_headers(original_sheet)
    # Create result workbook and sheet
    wb_result = Workbook()
    ws_result = wb_result.active
    # Create headers
    for cell in format:
        ws_result.cell(row=1, column=format.index(cell)+1, value=cell)
    #Initialize variables to save previous lat/lon
    prevlat = None
    prevlon = None
    prevtz = None
    # Create data into every row, one row at a time
    for i in range(2, original_sheet.max_row+1):
        #check if A_party
        A_party = original_sheet.cell(row=i, column=original_headers['A Number']).value == targetnumber
        column = 0
        # Create main data
        for cell in main_headers:
            column += 1
            realdate_tz = None
            if cell == 'A or B':
                continue
            elif cell in original_headers:
                ws_result.cell(row=i, column=column, value=original_sheet.cell(row=i, column=original_headers[cell]).value)
            elif cell == 'ID':
                ws_result.cell(row=i, column=column, value=i-1)
            elif cell == 'victimdatetime':
                #Victim's datetime
                origdate = original_sheet.cell(row=i, column=original_headers["UTC Date"]).value
                origtime = original_sheet.cell(row=i, column=original_headers["UTC Time"]).value
                origtimespecifics = origtime.split(":")
                x = 0
                while True:
                    if x == 3:
                        break
                    origtimespecifics[x] = int(origtimespecifics[x])
                    x += 1
                origtime_full = origdate + timedelta(hours=origtimespecifics[0], minutes=origtimespecifics[1], seconds=origtimespecifics[2])
                origtime_tz = alpha_tz.localize(origtime_full)
                realdate_tz = origtime_tz.astimezone(main_tz)
                realdate_tz = realdate_tz.replace(tzinfo=None)
                ws_result.cell(row=i, column=column, value=realdate_tz)
                #Target's datetime and timezone
                if A_party:
                    lat = original_sheet.cell(row=i, column=original_headers["A Start Location Latitude"]).value
                    lon = original_sheet.cell(row=i, column=original_headers["A Start Location Longitude"]).value
                else:
                    lat = original_sheet.cell(row=i, column=original_headers["B Start Location Latitude"]).value
                    lon = original_sheet.cell(row=i, column=original_headers["B Start Location Longitude"]).value
                # If coords same as previous, skip timezone calculation as it is extremely time intensive
                if lat == prevlat and lon == prevlon:
                    targettz = prevtz
                else:
                    targettz_raw = timezone_coord(lat, lon)
                    targettz = pytz.timezone(targettz_raw)
                targettime = origtime_tz.astimezone(targettz)
                targettime_fix = targettime.replace(tzinfo=None)
                ws_result.cell(row=i, column=column+1, value=targettime_fix)
                #Target timezone
                print(i, targettz_raw)
                prevlat = lat
                prevlon = lon
                prevtz = targettz
                ws_result.cell(row=i, column=column+2, value=targettz_raw)
            # elif cell == 'endtime':
            #     durationsec = original_sheet.cell(row=i, column=original_headers["Duration"]).value
            #     endtime = realdate_tz + timedelta(seconds=durationsec)
            #     ws_result.cell(row=i, column=column, value=endtime)
            # else: #expand with more elifs for specific cell types
            #     ws_result.cell(row=i, column=column, value='0')
        # Iterate over data that requires knowledge of whether A or B party.
        # If A party is our target, do this.
        if A_party:
            # Set target to A party in sheet
            ws_result.cell(row=i, column=column, value='A')
            # Iterate over the party-dependent data and write to sheet
            for cell in alphatel_A:
                column += 1
                if cell in original_headers:
                    # Check the temporary value if you want to do some specific modifications to it. In my case I want to check phone number to see if it's one of the targets
                    tempvalue = original_sheet.cell(row=i, column=original_headers[cell]).value
                    if tempvalue in target_numbers.values():
                        tempvalue = (list(target_numbers.keys())
                        [list(target_numbers.values()).index(tempvalue)])
                    ws_result.cell(row=i, column=column, value=tempvalue)
        # If A party is not our target, do this
        else:
            # Set target to B party in sheet
            ws_result.cell(row=i, column=column, value='B')
            # Iterate over party-dependent data but with reversed A and B party
            for cell in alphatel_B:
                column += 1
                if cell in original_headers:
                    # Check the temporary value if you want to do some specific modifications to it. In my case I want to check phone number to see if it's one of the targets
                    tempvalue = original_sheet.cell(row=i, column=original_headers[cell]).value
                    if tempvalue in target_numbers.values():
                        tempvalue = (list(target_numbers.keys())
                        [list(target_numbers.values()).index(tempvalue)])
                    ws_result.cell(row=i, column=column, value=tempvalue)
    return wb_result
