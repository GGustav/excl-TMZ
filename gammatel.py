import openpyxl
from openpyxl import Workbook
# from utility import target_providers, column_headers
# from datetime import datetime
# import pytz
from utility import *
import copy

# Sure wishing I had fuzzy search right about now to do this automatically from the sheet
gammatel_A = ['B Number', 'A Party Start Location Name',
              'B Party Start Location Name', 'A Party End Location Name', 'B Party End Location Name', 'A Party Start Base Station ID', 'B Party Start Base Station ID', 'A Party End Base Station ID', 'B Party End Base Station ID',
              'A Party Start Location Latitude', 'A Party Start Location Longitude', 'A Party Start Location Bearing', 'B Party Start Location Latitude', 'B Party Start Location Longitude', 'B Party Start Location Bearing',
              'A Party End Location Latitude', 'A Party End Location Longitude', 'A Party End Location Bearing', 'B Party End Location Latitude', 'B Party End Location Longitude', 'B Party End Location Bearing']

# Create gammatel_B automatically since I am not typing all that out again
gammatel_B = copy.deepcopy(gammatel_A)
gammatel_B[0] = 'A Party Number'
for i in range(1, len(gammatel_A)):
    if gammatel_B[i][0] == 'A':
        gammatel_B[i] = gammatel_B[i].replace('A', 'B', 1)
    elif gammatel_B[i][0] == 'B':
        gammatel_B[i] = gammatel_B[i].replace('B', 'A', 1)

gamma_tz = pytz.timezone("Etc/UTC")

def gamma_sheet(original_sheet, targetnumber):
    original_headers = column_headers(original_sheet)
    # Create result workbook and sheet
    wb_result = Workbook()
    ws_result = wb_result.active
    # Create headers
    for cell in format:
        ws_result.cell(row=1, column=format.index(cell)+1, value=cell)
    #Initialize variables to save previous lat/lon
    prevlat = None
    prevlon = None
    prevtz = None
    # Create data into every row, one row at a time
    for i in range(2, original_sheet.max_row+1):
        column = 0
        A_party = original_sheet.cell(row=i, column=original_headers['A Party Number']).value == targetnumber
        # Create main data
        for cell in main_headers:
            column += 1
            if cell == 'A or B':
                continue
            elif cell in original_headers:
                ws_result.cell(row=i, column=column, value=original_sheet.cell(row=i, column=original_headers[cell]).value)
            elif cell == 'ID':
                ws_result.cell(row=i, column=column, value=i - 1)
            elif cell == 'victimdatetime':
                origtime = original_sheet.cell(row=i, column=original_headers["UTC Time"]).value
                origtime_tz = gamma_tz.localize(origtime)
                realdate_tz = origtime_tz.astimezone(main_tz)
                realdate_tz = realdate_tz.replace(tzinfo=None)
                ws_result.cell(row=i, column=column, value=realdate_tz)
                #Target's datetime and timezone
                if A_party:
                    lat = original_sheet.cell(row=i, column=original_headers["A Party Start Location Latitude"]).value
                    lon = original_sheet.cell(row=i, column=original_headers["A Party Start Location Longitude"]).value
                else:
                    lat = original_sheet.cell(row=i, column=original_headers["B Party Start Location Latitude"]).value
                    lon = original_sheet.cell(row=i, column=original_headers["B Party Start Location Longitude"]).value
                # If coords same as previous, skip timezone calculation as it is extremely time intensive
                if lat == prevlat and lon == prevlon:
                    targettz = prevtz
                else:
                    targettz_raw = timezone_coord(lat, lon)
                    targettz = pytz.timezone(targettz_raw)
                targettime = origtime_tz.astimezone(targettz)
                targettime_fix = targettime.replace(tzinfo=None)
                ws_result.cell(row=i, column=column+1, value=targettime_fix)
                #Target timezone
                print(targettime)
                print(i, targettz_raw)
                prevlat = lat
                prevlon = lon
                prevtz = targettz
                ws_result.cell(row=i, column=column+2, value=targettz_raw)
            # else:  # expand with more elifs for specific cell types
            #     ws_result.cell(row=i, column=column, value='0')
                # Iterate over data that requires knowledge of whether A or B party.
        # If A party is our target, do this.
        if A_party:
            # Set target to A party in sheet
            ws_result.cell(row=i, column=column, value='A')

            # Iterate over the party-dependent data and write to sheet
            for cell in gammatel_A:
                column += 1
                if cell in original_headers:
                    # Check the temporary value if you want to do some specific modifications to it. In my case I want to check phone number to see if it's one of the targets
                    tempvalue = original_sheet.cell(row=i, column=original_headers[cell]).value
                    if tempvalue in target_numbers.values():
                        tempvalue = (list(target_numbers.keys())
      [list(target_numbers.values()).index(tempvalue)])
                    ws_result.cell(row=i, column=column, value=tempvalue)
        # If A party is not our target, do this
        else:
            # Set target to B party in sheet
            ws_result.cell(row=i, column=column, value='B')
            # Iterate over party-dependent data but with reversed A and B party
            for cell in gammatel_B:
                column += 1
                if cell in original_headers:
                    # Check the temporary value if you want to do some specific modifications to it. In my case I want to check phone number to see if it's one of the targets
                    tempvalue = original_sheet.cell(row=i, column=original_headers[cell]).value
                    if tempvalue in target_numbers.values():
                        tempvalue = (list(target_numbers.keys())
                        [list(target_numbers.values()).index(tempvalue)])
                    ws_result.cell(row=i, column=column, value=tempvalue)
    return wb_result
